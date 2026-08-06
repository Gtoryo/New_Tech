"""
=============================================================================
GÉNÉRATEUR DE DONNÉES HISTORIQUES — PME MULTISERVICE POINTE-NOIRE, CONGO
=============================================================================
Simule 2 ans d'activité (01/06/2024 → 01/06/2026) avec anomalies réalistes
pour tester un pipeline ETL et un modèle Prophet de prévision de demande.

Fichiers générés :
  - ventes_historiques.xlsx
  - depenses_et_achats.xlsx
  - suivi_clients_prospects.xlsx

Les anomalies de saisie réinjectées (formats de dates hétérogènes, fautes de
frappe sur les libellés, doublons de factures, montants sentinelles et négatifs,
variantes de noms de villes) reproduisent les familles de défauts relevées lors
du cadrage métier avec la Gestionnaire.

Usage : exécuter depuis la racine du projet.
    python generated_data/generate_data.py

Les graines aléatoires sont fixes : deux exécutions produisent des fichiers
strictement identiques, ce qui rend l'ensemble du pipeline reproductible.
=============================================================================
"""

import os
import random
import sys
from datetime import date, timedelta

import numpy as np
import pandas as pd

# La console Windows utilise cp1252 par défaut, qui ne sait pas encoder les
# caractères Unicode des messages de progression. Sans cette ligne, le script
# échoue sur un poste Windows avec UnicodeEncodeError.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

# Graines fixes : garantissent que le jeu de travail est reproductible à
# l'identique, y compris les anomalies injectées.
np.random.seed(42)
random.seed(42)

# Chemin relatif à la racine du projet : le script doit être lancé depuis
# celle-ci (voir README, section « Reproduire le jeu de travail »).
OUTPUT_DIR = "data"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─────────────────────────────────────────────────────────────────────────────
# DONNÉES DE RÉFÉRENCE
# ─────────────────────────────────────────────────────────────────────────────

EMPLOYES = {
    "Sérigraphie":      ["Christian Loubaki", "Grâce Mavoungou"],
    "Imprimerie":       ["Prince Bouanga", "Divine Ngoma"],
    "Vidéosurveillance":["Rodrigue Massamba", "Sylvain Moukengue"],
    "Maintenance":      ["Rodrigue Massamba", "Sylvain Moukengue",
                         "Franck Elenga", "Papy Nzouzi"],
}

# Clients locaux réalistes (Pointe-Noire)
CLIENTS = [
    ("Pharmacie Sainte-Marie",      "06 650 1234"), ("École Privée Les Aiglons",   "06 712 5678"),
    ("Salon Beauté Nathalie",        "05 530 9876"), ("Imprimerie Riviera Plus",    None),
    ("Restaurant Le Ngombe",         "06 840 3311"), ("Boutique Mode Africaine",    "05 621 7788"),
    ("Association AFRIJEP",          "06 900 4455"), ("Église Évangélique Centrale","05 710 2233"),
    ("Collège Technique Mvou-Mvou",  "06 622 0011"), ("Mairie du 1er Arr.",         "06 531 8800"),
    ("Cabinet Médical Dr. Mabiala",  "05 840 6677"), ("Auto-école Étoile",          "06 712 9900"),
    ("ONG Femmes et Développement",  "06 611 3344"), ("Librairie Lecture Facile",   "05 530 1122"),
    ("Snack-Bar Chez Tantine",       None),          ("Garage Mécanique Bolloré",   "06 700 5566"),
    ("Centre de Santé Loandjili",    "05 841 2233"), ("Hôtel Résidence du Fleuve",  "06 900 7788"),
    ("COGECO (Coopérative locale)",  "06 622 4400"), ("École Primaire Pointe-Noire","06 513 0099"),
    ("Banque Populaire du Congo",    "06 730 1122"), ("Supermarché Mbota",          "05 620 9988"),
    ("Menuiserie Frères Nianga",     None),          ("Studio Photo Lumière",       "06 841 5544"),
    ("Quincaillerie du Centre",      "05 712 3388"),
]

# ─────────────────────────────────────────────────────────────────────────────
# CATALOGUE PRODUITS / SERVICES (prix en FCFA, réalistes Congo)
# ─────────────────────────────────────────────────────────────────────────────

CATALOGUE = {
    "serigraphie": [
        ("Sérigraphie", "T-shirt personnalisé (dotation entreprise)", 3500,  5000,  20, 200),
        ("Sérigraphie", "T-shirt personnalisé (deuil)",               2500,  3500,  10,  80),
        ("Sérigraphie", "T-shirt personnalisé (événement)",           3000,  4500,  15, 150),
        ("Sérigraphie", "Casquette brodée",                           4000,  6500,   5,  50),
        ("Sérigraphie", "Polo personnalisé",                          5500,  8000,  10, 100),
    ],
    "imprimerie": [
        ("Imprimerie", "Flyers A5 (500 exemplaires)",  15000, 25000,  1,  10),
        ("Imprimerie", "Flyers A4 (1000 exemplaires)", 25000, 40000,  1,   5),
        ("Imprimerie", "Cartes de visite (100 pcs)",    8000, 15000,  1,  20),
        ("Imprimerie", "Banderole 1m x 3m",            20000, 35000,  1,   5),
        ("Imprimerie", "Catalogue produits (A4, 8p)",  35000, 60000,  1,   3),
        ("Imprimerie", "Affiche A3 plastifiée",         3500,  6000,  5,  50),
    ],
    "video": [
        ("Vidéosurveillance", "Installation caméra intérieure (unité)", 45000,  75000, 1, 4),
        ("Vidéosurveillance", "Installation caméra extérieure (unité)", 55000,  90000, 1, 4),
        ("Vidéosurveillance", "Kit vidéosurveillance 4 caméras",       280000, 450000, 1, 1),
        ("Vidéosurveillance", "Maintenance système CCTV",               25000,  50000, 1, 2),
    ],
    "maintenance": [
        ("Maintenance", "Réparation PC portable",     25000, 60000, 1, 2),
        ("Maintenance", "Réparation PC bureau",       20000, 50000, 1, 3),
        ("Maintenance", "Installation logiciel/OS",   15000, 30000, 1, 3),
        ("Maintenance", "Récupération de données",    30000, 80000, 1, 1),
        ("Maintenance", "Maintenance réseau LAN",     40000, 80000, 1, 1),
    ],
}

# ─────────────────────────────────────────────────────────────────────────────
# UTILITAIRES
# ─────────────────────────────────────────────────────────────────────────────

def date_range_list(start: date, end: date):
    delta = (end - start).days
    return [start + timedelta(days=i) for i in range(delta + 1)]

def saisonnalite_poids(d: date) -> float:
    m = d.month
    if m == 12:              return 2.2
    elif m in (9, 10):       return 1.8
    elif m == 8 and d.day >= 10 and d.day <= 20: return 1.4
    elif m in (7, 8):        return 0.85
    elif m in (1, 2):        return 0.90
    else:                    return 1.0

def gen_facture_id(d: date, idx: int) -> str:
    return f"FAC-{d.strftime('%Y%m')}-{idx:04d}"

FAUTES_SERVICE = {
    "Sérigraphie":      ["Sérigrafe", "Serigraphy", "Sérigraphie ", "sérigrpahie", "Sérigr."],
    "Imprimerie":       ["Imprimerie ", "imprimrie", "Imprimeri", "IMPRIMERIE"],
    "Vidéosurveillance":["Video surveillance", "Vidéo-surveillance", "Vidéosur."],
    "Maintenance":      ["Reparation", "Maintenace", "Maintenance ", "maint."],
}

def service_avec_faute(service: str) -> str:
    if random.random() < 0.07:
        return random.choice(FAUTES_SERVICE.get(service, [service]))
    return service

def format_date_aleatoire(d: date) -> str:
    r = random.random()
    if r < 0.72:   return d.strftime("%d/%m/%Y")
    elif r < 0.85: return d.strftime("%Y-%m-%d")
    elif r < 0.93: return d.strftime("%d-%m-%Y")
    else:
        mois_fr = ["","Janvier","Février","Mars","Avril","Mai","Juin",
                   "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
        return f"Début {mois_fr[d.month]} {d.year}"

# ─────────────────────────────────────────────────────────────────────────────
# FICHIER 1 : VENTES HISTORIQUES
# ─────────────────────────────────────────────────────────────────────────────

def generer_ventes() -> pd.DataFrame:
    print("  → Génération des ventes historiques...")
    start = date(2024, 6, 1)
    end   = date(2026, 6, 1)
    all_dates = date_range_list(start, end)

    rows = []
    facture_counter = 1

    for d in all_dates:
        poids = saisonnalite_poids(d)
        n_ventes = np.random.poisson(lam=3.5 * poids)
        if d.weekday() == 6:
            n_ventes = max(0, n_ventes - 2)

        for _ in range(n_ventes):
            secteur = random.choices(
                ["serigraphie", "imprimerie", "video", "maintenance"],
                weights=[35, 35, 15, 15]
            )[0]

            produit = random.choice(CATALOGUE[secteur])
            service_type, description, pu_min, pu_max, qte_min, qte_max = produit

            client, telephone = random.choice(CLIENTS)
            quantite   = random.randint(qte_min, qte_max)
            prix_unit  = round(random.randint(pu_min, pu_max) / 100) * 100
            total_calc = quantite * prix_unit

            if secteur == "video":           cle_employe = "Vidéosurveillance"
            elif secteur == "maintenance":   cle_employe = "Maintenance"
            elif secteur == "serigraphie":   cle_employe = "Sérigraphie"
            else:                            cle_employe = "Imprimerie"
            employe = random.choice(EMPLOYES[cle_employe])

            statut = random.choices(
                ["Payé", "Payé", "Payé", "En attente", "Partiel"],
                weights=[65, 65, 65, 20, 15]
            )[0]

            rows.append({
                "Date":              format_date_aleatoire(d),
                "Facture_ID":        gen_facture_id(d, facture_counter),
                "Client":            client,
                "Telephone":         telephone,
                "Service_Type":      service_avec_faute(service_type),
                "Description":       description,
                "Quantite":          quantite,
                "Prix_Unitaire":     prix_unit,
                "Total":             total_calc,
                "Employe_En_Charge": employe,
                "Statut_Paiement":   statut,
                "_date_obj":         d,
            })
            facture_counter += 1

    df = pd.DataFrame(rows)
    n  = len(df)

    mask_tel = np.random.random(n) < 0.08
    df.loc[mask_tel, "Telephone"] = np.nan
    mask_emp = np.random.random(n) < 0.04
    df.loc[mask_emp, "Employe_En_Charge"] = np.nan
    mask_pu = np.random.random(n) < 0.03
    df.loc[mask_pu, "Prix_Unitaire"] = np.nan
    mask_incoh = np.random.random(n) < 0.04
    df.loc[mask_incoh, "Total"] = random.choice([0, np.nan, 999])

    n_doublons = int(n * 0.05)
    idx_doublons = np.random.choice(df.index, size=n_doublons, replace=False)
    doublons_df  = df.loc[idx_doublons].copy()
    df = pd.concat([df, doublons_df], ignore_index=True)
    df = df.sample(frac=1, random_state=42).reset_index(drop=True)
    df.drop(columns=["_date_obj"], inplace=True)

    print(f"     ✓ {len(df)} lignes générées (dont ~{n_doublons} doublons)")
    return df


# ─────────────────────────────────────────────────────────────────────────────
# FICHIER 2 : DÉPENSES ET ACHATS
# ─────────────────────────────────────────────────────────────────────────────

FOURNISSEURS = [
    "Quincaillerie du Centre", "Grossiste Grand Marché", "Bureau Top",
    "Impex Congo SARL", "Papeterie Moderne PNR", "Couleurs & Co.",
    "Fournisseur Textile Makola",
]

ARTICLES_BASE = {
    "Encre": [
        ("Encre sérigraphie noir 1L",   8500,  14000),
        ("Encre sérigraphie couleur 1L", 10000, 18000),
        ("Encre offset noire (pot 5kg)", 35000, 55000),
        ("Encre UV flexo",              12000, 20000),
    ],
    "T-shirt vierge": [
        ("T-shirt blanc 100% coton",  1200, 2000),
        ("T-shirt couleur S-M-L",     1500, 2500),
        ("Polo blanc unisexe",        2500, 4000),
    ],
    "Papier": [
        ("Ramette papier A4 80g",     3500,  5500),
        ("Papier couché brillant A3", 8000, 14000),
        ("Papier autocollant blanc",  6000, 11000),
        ("Carton d'impression 250g",  9000, 16000),
    ],
    "Autre": [
        ("Toile polyester (mètre)",   2500,  4500),
        ("Film plastique thermorétractable", 15000, 25000),
        ("Cartouche imprimante HP",   18000, 30000),
        ("Toner laser noir",          25000, 45000),
        ("Alcool isopropylique 1L",    4000,  7000),
    ],
}

VARIANTES_ORTHO = {
    "T-shirt blanc 100% coton": ["T-shirt Blanc", "Tshirt blanc", "TSHIRT BLANC", "t-shirt blanc coton"],
    "Ramette papier A4 80g":    ["Ramette A4", "Papier A4 80g", "RAMETTE A4 80G", "ramette papier"],
    "Encre sérigraphie noir 1L":["Encre Noire Séri", "encre sérigraphie noire", "ENCRE SERI NOIR"],
}

def variante_article(nom: str) -> str:
    if nom in VARIANTES_ORTHO and random.random() < 0.25:
        return random.choice(VARIANTES_ORTHO[nom])
    return nom

def generer_depenses() -> pd.DataFrame:
    print("  → Génération des dépenses et achats...")
    start = date(2024, 6, 1)
    end   = date(2026, 6, 1)
    all_dates = date_range_list(start, end)

    rows = []
    for d in all_dates:
        if random.random() > 0.30:
            continue
        n_achats = random.randint(1, 3)
        for _ in range(n_achats):
            categorie = random.choices(
                list(ARTICLES_BASE.keys()), weights=[30, 25, 30, 15]
            )[0]
            article_info = random.choice(ARTICLES_BASE[categorie])
            nom_article, px_min, px_max = article_info

            quantite     = random.randint(1, 20)
            prix_total   = round(random.randint(px_min * quantite, px_max * quantite) / 500) * 500
            mode_paiement = random.choice(["Espèces", "Espèces", "Mobile Money", "Airtel Money"])

            rows.append({
                "Date":             format_date_aleatoire(d),
                "Fournisseur":      random.choice(FOURNISSEURS),
                "Article":          variante_article(nom_article),
                "Categorie":        categorie,
                "Quantite":         quantite,
                "Prix_Achat_Total": prix_total,
                "Mode_Paiement":    mode_paiement,
            })

    df = pd.DataFrame(rows)
    n  = len(df)

    idx_vides = np.random.choice(df.index, size=int(n * 0.02), replace=False)
    df.loc[idx_vides, :] = np.nan
    idx_neg = np.random.choice(df.index, size=int(n * 0.02), replace=False)
    df.loc[idx_neg, "Prix_Achat_Total"] = df.loc[idx_neg, "Prix_Achat_Total"].apply(
        lambda x: -abs(x) if pd.notna(x) else x
    )

    print(f"     ✓ {n} lignes générées")
    return df


# ─────────────────────────────────────────────────────────────────────────────
# FICHIER 3 : SUIVI CLIENTS / PROSPECTS
# ─────────────────────────────────────────────────────────────────────────────

VARIANTES_VILLE_PN = [
    "Pointe-Noire", "Pointe-Noire", "Pointe-Noire",
    "Pointe Noire", "PN", "Pte-Noire", "pointenoire",
    "POINTE-NOIRE", "Pointe-noire", "P.Noire",
]

PROSPECTS_SUPPLEMENTAIRES = [
    ("Jean-Michel Nzabi",    "Auto-école Étoile",          "06 712 9900", "jmnzabi@gmail.com"),
    ("Rosalie Mouanga",      "Association AFRIJEP",        "06 900 4455", None),
    ("Théodore Kiabouanga",  "Snack-Bar Chez Tantine",     None,          "theokia@yahoo.fr"),
    ("Carine Bouesso",       "Studio Photo Lumière",       "06 841 5544", "cbouesso@outlook.com"),
    ("Fernand Loemba",       "Menuiserie Frères Nianga",   None,          None),
    ("Annick Pambou",        "Pharmacie Sainte-Marie",     "06 650 1234", "apambou@pharmasante.cg"),
    ("Patrick Madzimba",     "",                           "05 712 0011", None),
    ("Ghislaine Koubemba",   "ONG Femmes et Développement","06 611 3344", "gkouemba@ong-fd.org"),
    ("Sylvestre Makaya",     "Librairie Lecture Facile",   "05 530 1122", None),
    ("Nadège Bitsindou",     "Hôtel Résidence du Fleuve",  "06 900 7788", "nadege@residencefleuve.cg"),
    ("Elie Mpouya",          "Supermarché Mbota",          "05 620 9988", "empouya@mbota.cg"),
    ("Véronique Nzoussi",    "École Primaire Pointe-Noire","06 513 0099", None),
    ("Cédric Louboto",       "Banque Populaire du Congo",  "06 730 1122", "clouboto@bpc.cg"),
    ("Martine Banzouzi",     "COGECO",                     "06 622 4400", None),
    ("Arnold Loutaya",       "Centre de Santé Loandjili",  "05 841 2233", "aloutaya@csl.cg"),
    ("Serge Ondongo",        "Groupe Ondongo Brazza",      "06 500 8811", "sondongo@gmail.com"),
    ("Laurence Kimpioka",    "Cabinet Comptable BZV",      "06 501 7722", None),
    ("Bertrand Nkounkou",    "Imprimerie Centrale BZV",    "06 502 6633", "bnkounkou@icbzv.cg"),
]

NOTES_POSSIBLES = [
    "Client fidèle, toujours ponctuel",
    "À relancer pour commande annuelle t-shirts",
    "Intéressé par pack flyers + cartes de visite",
    "Payement souvent en retard",
    "Contact via recommandation de Mme Tchicaya",
    "Demande devis banderole — en attente",
    "A commandé en décembre, à fidéliser",
    np.nan, np.nan, np.nan,
    "Prospect chaud — à rappeler",
    "RDV fixé mais annulé, recontacter",
    "Gros volume potentiel (dotation annuelle)",
]

def generer_clients() -> pd.DataFrame:
    print("  → Génération du suivi clients/prospects...")
    rows = []

    for client, telephone in CLIENTS:
        ville_r = random.random()
        ville = random.choice(VARIANTES_VILLE_PN) if ville_r < 0.88 else "Brazzaville"
        rows.append({
            "Nom_Client":  client,
            "Entreprise":  client,
            "Telephone":   telephone if pd.notna(telephone) else np.nan,
            "Email":       f"{client.lower().replace(' ', '.').replace('é','e').replace('è','e')[:15]}@gmail.com"
                           if random.random() > 0.4 else np.nan,
            "Ville":       ville,
            "Notes":       random.choice(NOTES_POSSIBLES),
        })

    for nom, entreprise, tel, email in PROSPECTS_SUPPLEMENTAIRES:
        if nom in ("Serge Ondongo", "Laurence Kimpioka", "Bertrand Nkounkou"):
            ville = "Brazzaville"
        else:
            ville = random.choice(VARIANTES_VILLE_PN) if random.random() < 0.88 else "Brazzaville"
        rows.append({
            "Nom_Client":  nom,
            "Entreprise":  entreprise if entreprise else np.nan,
            "Telephone":   tel if tel else np.nan,
            "Email":       email if email else np.nan,
            "Ville":       ville,
            "Notes":       random.choice(NOTES_POSSIBLES),
        })

    df = pd.DataFrame(rows)
    n_doublons = max(2, int(len(df) * 0.10))
    idx_dup    = np.random.choice(df.index, size=n_doublons, replace=False)
    dup_df     = df.loc[idx_dup].copy()
    dup_df["Nom_Client"] = dup_df["Nom_Client"].apply(
        lambda x: x.upper() if isinstance(x, str) else x
    )
    df = pd.concat([df, dup_df], ignore_index=True)
    df = df.sample(frac=1, random_state=7).reset_index(drop=True)

    print(f"     ✓ {len(df)} contacts générés")
    return df


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def sauvegarder_excel(df: pd.DataFrame, nom_fichier: str, nom_onglet: str):
    path = os.path.join(OUTPUT_DIR, nom_fichier)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=nom_onglet, index=False)
        ws = writer.sheets[nom_onglet]
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
        for cell in ws[1]:
            cell.font      = Font(bold=True, name="Arial", size=10)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial", size=9)
    print(f"     ✓ Sauvegardé → {path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n" + "="*65)
    print("  GÉNÉRATEUR DE DONNÉES — PME MULTISERVICE POINTE-NOIRE")
    print("="*65)

    print("\n[1/3] ventes_historiques.xlsx")
    df_ventes = generer_ventes()
    sauvegarder_excel(df_ventes, "ventes_historiques.xlsx", "Ventes")

    print("\n[2/3] depenses_et_achats.xlsx")
    df_depenses = generer_depenses()
    sauvegarder_excel(df_depenses, "depenses_et_achats.xlsx", "Dépenses")

    print("\n[3/3] suivi_clients_prospects.xlsx")
    df_clients = generer_clients()
    sauvegarder_excel(df_clients, "suivi_clients_prospects.xlsx", "Clients")

    print("\n" + "="*65)
    print("  ✅ GÉNÉRATION TERMINÉE")
    print("="*65)
